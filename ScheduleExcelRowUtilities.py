# ******************************************************
#
# Southeastern Railway Museum, Project Task Scheduler
#
# Use Python 3.5+ or better
#
# Author/Owner: C. Hardt
#
# ******************************************************

import openpyxl

# class provides utilites to quickly locate excel rows, and take actions on them
class ExcelRowUtil:

    def __init__(self):
        print("Excel Row Class Created")

    # ***************************************
    #
    # InsertRow()
    #
    # Inserts a blank row in the row number specified
    # Move the current row down one row
    #
    # ***************************************
    def InsertRow(WorksheetName, RowNumberToInsert):

        try:

            max_row = WorksheetName.max_row
            max_col = WorksheetName.max_column

            # Now do the rest of it. Note the row offset.
            for iRowNumber in range(max_row, RowNumberToInsert, -1):
                for col_num in range (1, max_col):
                    WorksheetName.cell(iRowNumber + 1, col_num).value = WorksheetName.cell(iRowNumber, col_num).value
                    print (WorksheetName.cell(iRowNumber + 1, col_num).value)

            # Blank out all the cells in the row we just inserted
            for col_num in range (1, max_col):
                WorksheetName.cell(RowNumberToInsert, col_num).value = ""
        except:
            print("Exception in InsertRow()")

        return

    # ***************************************
    #
    # DeleteRow()
    #
    # Deletes the specifed row number from the sheet
    # Move the row below up one row
    #
    # ***************************************
    def DeleteRow(WorksheetName, RowNumberToDelete):

        try:
            max_row = WorksheetName.max_row
            max_col = WorksheetName.max_column

            # Now do the rest of it. Note the row offset.
            for iRowNumber in range(RowNumberToDelete, max_row):
                for col_num in range (1, max_col):
                    WorksheetName.cell(iRowNumber, col_num).value = WorksheetName.cell(iRowNumber + 1, col_num).value
                    print (WorksheetName.cell(iRowNumber, col_num).value)

        except:
            print("Exception in DeleteRow()")

        return

    # ***************************************
    #
    # CopyRow()
    #
    # Copies the specifed row number to the
    # specified row number.  It will overwrite the from row.
    #
    # ***************************************
    def CopyRow (WorksheetName, RowNumberToCopyFrom, RowNumberToCopyTo):

        try:

            max_row = WorksheetName.max_row
            max_col = WorksheetName.max_column

            for col_num in range(1, max_col):
                WorksheetName.cell(RowNumberToCopyTo, col_num).value = WorksheetName.cell(RowNumberToCopyFrom, col_num).value
                print("Row: " + str(RowNumberToCopyFrom) + "Copied to row: " + str(RowNumberToCopyTo))

        except:
            print("Exception in CopyRow()")

        return

    # ***************************************
    #
    # SwapRow()
    #
    # Swaps the position of the two specified rows.  This
    # is used during a sheet sort operation
    #
    # ***************************************
    def SwapRows (WorksheetName, RowNumberToCopyFrom, RowNumberToCopyTo):

        try:

            max_row = WorksheetName.max_row
            max_col = WorksheetName.max_column

            # make sure our arguments are in range, if they are, then proceed, else exception
            if (RowNumberToCopyFrom >= 1 and RowNumberToCopyFrom <= max_row) and \
                    (RowNumberToCopyTo >= 1 and RowNumberToCopyTo <= max_row):

                # we are going to copy each column one by one, swapping the values as we go along
                for col_num in range(1, max_col):
                    TempValue = WorksheetName.cell(RowNumberToCopyTo, col_num).value
                    WorksheetName.cell(RowNumberToCopyTo, col_num).value = WorksheetName.cell(RowNumberToCopyFrom, col_num).value
                    WorksheetName.cell(RowNumberToCopyFrom, col_num).value = TempValue

                #print("Row: " + str(RowNumberToCopyFrom) + " Swapped with row: " + str(RowNumberToCopyTo))
            else:
                raise ValueError("SwapRows() - Row number(s) passed is not within acceptable range")
        except:
            print("Exception in SwapRows()")


        return