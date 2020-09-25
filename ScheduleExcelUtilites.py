# ******************************************************
#
# Southeastern Railway Museum, Project Task Scheduler
#
# Use Python 3.5+ or better
#
# Author/Owner: C. Hardt
#
# ******************************************************

import openpyxl
import sys

# Excel column related utilities
class ExcelColumnUtil:

    def __init__(self):
        print("Excel Column Class Created")
	
	# Return (Find) the column number with the associated column header 
    # Row in which the header can be located is the RowNumber
    def ExcelFindColumnHeader(WorkSheetName, sColumnName, iRowNumber):

        iColumnIndexNumber = -1  # we are going to set the column number to an error value just in case we throw an exception

        try:

            bFoundColumnFlag = False
            iMaxColumnNumber = WorkSheetName.max_column
            iMaxRowNumber = WorkSheetName.max_row

            # We will walk across all the columns (left to right), looking for a matching name
            for iColumnIndexNumber in range(1, iMaxColumnNumber, 1):

                sColumnValue = WorkSheetName.cell(iRowNumber, iColumnIndexNumber).value

                # Did we find a matching column header name?  If so we are done
                if (sColumnValue == sColumnName):
                    #print("Found " + sColumnName + " in Column " + str(iColumnIndexNumber))
                    bFoundColumnFlag = True
                    break

            #  we did not find the column we were told to look for, so print an error message
            if (bFoundColumnFlag == False):
                print("ERROR Column " + sColumnName + " Not Found!")

        except Exception as e:
            print("Exception in ExcelFindColumnHeader()" + str(e))
            print("Worksheet: " + WorkSheetName + " ColumnName: " + sColumnName + " Index: " + str(iColumnIndexNumber))



        return (iColumnIndexNumber)

