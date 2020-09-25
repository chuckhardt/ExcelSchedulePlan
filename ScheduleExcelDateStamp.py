# ******************************************************
#
# Southeastern Railway Museum, Project Task Scheduler
#
# Use Python 3.5+ or better
#
# Author/Owner: C. Hardt
#
# ******************************************************

import openpyxl
import datetime

from openpyxl.styles import Color, Fill
from openpyxl.styles import Font

from openpyxl.cell import Cell

from ScheduleExcelConvertDate import ConvertExcelDate

# ***************************************
#
# ExcelUpdateDateStamps()
#
# Convert the date used by Excel to a number we can work with
#
# ***************************************
def ExcelUpdateDateStamps(WorksheetName):

    #print("Scanning Worksheet Date Stamps.. " + WorksheetName)

    try:

        iMaxRowNumber = WorksheetName.max_row
        iMaxColumnNumber = WorksheetName.max_column

        # Now do the rest of it. Note the row offset.
        for iRowNumber in range(2, iMaxRowNumber + 1, 1):

            #  Convert the date used by Excel to a number we can work with
            ExcelTaskStartDate = ConvertExcelDate(WorksheetName.cell(iRowNumber, 7).value)

            # time estimate to complete the task/defect
            ExcelTaskOriginalEstimate = WorksheetName.cell(iRowNumber, 6).value

            OriginalEstimateConvertedToDays = ExcelTaskOriginalEstimate//28800
            OriginalEstimateConvertedToSeconds = ExcelTaskOriginalEstimate % 28800

            ExcelCalculatedTaskCompletionDate = ExcelTaskStartDate + datetime.timedelta(days=OriginalEstimateConvertedToDays, seconds=OriginalEstimateConvertedToSeconds)

            myCell = WorksheetName['H2']
            myCell.value = "12/03/19"
            myCell._set_number_format('MM/DD/YY HH:MM')
            myCell.style.font.name = 'Arial'
            myCell.style.number_format.format_code = 'MM/DD/YY HH:MM'
            WorksheetName.cell(iRowNumber, 8).value = ExcelCalculatedTaskCompletionDate

            print("Row " + str(iRowNumber) + \
                  " - Init Date: " + str(WorksheetName.cell(iRowNumber, 7).value) + \
                  " - Orig Est: " + str(ExcelTaskOriginalEstimate) + \
                  " - Est Days: " + str(OriginalEstimateConvertedToDays) + \
                  " Est Seconds: " + str(OriginalEstimateConvertedToSeconds) + \
                  " - Task Compl Date: " + str(ExcelCalculatedTaskCompletionDate))

    except Exception as e:
        print("Exception in ExcelUpdateDateStamps()" + str(e))

    return
