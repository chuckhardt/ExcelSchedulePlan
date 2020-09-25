# ******************************************************
#
# Southeastern Railway Museum, Project Task Scheduler
#
# Use Python 3.5+ or better
#
# Author/Owner: C. Hardt
#
# ******************************************************

import openpyxl

from ScheduleExcelUtilites import ExcelColumnUtil
from ScheduleExcelRowUtilities import ExcelRowUtil

# ***************************************
#
# ScheduleExcelSortSheet()
#
# We are going to sort the worksheet by work blocks first,
# with the lowest numbered blocks at the top of the sheet
#
# ***************************************

def ScheduleExcelSortSheet(WorksheetName):

    print("ScheduleExcelSortSheet(): Worksheet name: " + WorksheetName)

    try:

        iMaxColumnNumber = WorksheetName.max_column
        iMaxRowNumber = WorksheetName.max_row

        bDidASwapOperationOccur = True
        iSortLoopCount = 0
        iSwapCountPerPass = 0

        # create our excel column utilites class obj
        ExcelColumnUtilites = ExcelColumnUtil()

        kWorkBlkNumberColumn = ExcelColumnUtilites.ExcelFindColumnHeader(WorksheetName, "WorkBlk", 1) #7
        kAssigneeNameColumn = ExcelColumnUtilites.ExcelFindColumnHeader(WorksheetName, "Assignee", 1) #5
        kTaskNumberColumn = ExcelColumnUtilites.ExcelFindColumnHeader(WorksheetName, "Issue key", 1) #1

        # we will stay in this loop until there is nothing left to sort
        while (bDidASwapOperationOccur == True):

            bDidASwapOperationOccur = False
            iSortLoopCount += 1
            iSwapCountPerPass = 0

            ExcelRowUtilities = ExcelRowUtil()

            # We will start by sorting the sheet by WorkBlk numbers (lowest number first)
            for iRowIndexNumber in range(2, iMaxRowNumber, 1):

                # read in the WorkBlk name from the current line, verify that the field is not blank, correct if needed
                sWorkBlkCurrentLine = WorksheetName.cell(iRowIndexNumber, kWorkBlkNumberColumn).value
                if not sWorkBlkCurrentLine:
                    sWorkBlkCurrentLine = ""

                # read in the WorkBlk name from the next line, verify that the field is not blank, correct if needed
                sWorkBlkNextLine = WorksheetName.cell(iRowIndexNumber + 1, kWorkBlkNumberColumn).value
                if not sWorkBlkNextLine:
                    sWorkBlkNextLine = ""

                # does the WorkBlk number in current line >, < or equal to the next line?
                if sWorkBlkCurrentLine > sWorkBlkNextLine:
                    bDidASwapOperationOccur = True
                    iSwapCountPerPass += 1
                    ExcelRowUtilities.SwapRows(WorksheetName, iRowIndexNumber + 1, iRowIndexNumber)

                # ok, looks like the WorkBlk numbers between the two lines match, so let's see if the resource names need sorting
                elif sWorkBlkCurrentLine == sWorkBlkNextLine:

                    # now read in the resource assignee name from the current line, correct if an empty field
                    sAssigneeCurrentLine = WorksheetName.cell(iRowIndexNumber, kAssigneeNameColumn).value
                    if not sAssigneeCurrentLine:
                        sAssigneeCurrentLine = ""

                    # now read in the resource assignee name from the next line, correct if an empty field
                    sAssigneeNextLine = WorksheetName.cell(iRowIndexNumber + 1, kAssigneeNameColumn).value
                    if not sAssigneeNextLine:
                        sAssigneeNextLine = ""

                    # Is the Assignee Name in current line >, < or equal to the next line?
                    if sAssigneeCurrentLine.lower() > sAssigneeNextLine.lower():
                        bDidASwapOperationOccur = True
                        iSwapCountPerPass += 1
                        ExcelRowUtilities.SwapRows(WorksheetName, iRowIndexNumber + 1, iRowIndexNumber)
                        print ("Moving row based on assignee: " + sAssigneeCurrentLine + " " + sAssigneeNextLine )

                    # if the assignee names match, now sort on the task numbers
                    elif sAssigneeCurrentLine.lower() == sAssigneeNextLine.lower():

                        # now read in the task ticket number from the current line, correct if an empty field
                        sTaskNumberNumberCurrentLine = WorksheetName.cell(iRowIndexNumber, kTaskNumberColumn).value
                        if not sTaskNumberNumberCurrentLine:
                            sTaskNumberNumberCurrentLine = ""

                        # now read in the task number from the next line, correct if an empty field
                        sTaskNumberNumberNextLine = WorksheetName.cell(iRowIndexNumber + 1, kTaskNumberColumn).value
                        if not sTaskNumberNumberNextLine:
                            sTaskNumberNumberNextLine = ""

                        # need to convert the task ticket to a number we can sort on.
                        # we are trim the task string number, then convert to an int
                        iTaskNumberNumberCurrentLine = int(sTaskNumberNumberCurrentLine[3:5])
                        iTaskNumberNumberNextLine = int(sTaskNumberNumberNextLine[3:5])

                        # Is the ticket number  in current line >, < or equal to the next line?
                        if iTaskNumberNumberCurrentLine > iTaskNumberNumberNextLine:
                            bDidASwapOperationOccur = True
                            iSwapCountPerPass += 1
                            ExcelRowUtilities.SwapRows(WorksheetName, iRowIndexNumber + 1, iRowIndexNumber)
                            print("Moving row based on task number: " + sAssigneeCurrentLine + " " + sAssigneeNextLine)

            print("Sorting, Pass: " + str(iSortLoopCount) + " Changes Made: " + str(iSwapCountPerPass))

    except Exception as e:
        print("Exception in ScheduleExcelSortSheet()" + str(e))

    return

# ***************************************
#
# ScheduleExcelFormatDate()
#
# Need to firmate the date in a consistent format
#
# ***************************************
def ScheduleExcelFormatDate(WorksheetName):

    iMaxRowNumber = WorksheetName.max_row

    ExcelColumnUtilites = ExcelColumnUtil()
    kDateColumn = ExcelColumnUtilites.ExcelFindColumnHeader(WorksheetName, "Original Estimate", 1)

    # We will start by walking thru the spreadsheet, and converting the date from seconds to days
    for iRowIndexNumber in range(2, iMaxRowNumber + 1, 1):

        # read in the WorkBlk name from the current line, verify that the field is not blank, correct if needed
        DateInSeconds = WorksheetName.cell(iRowIndexNumber, kDateColumn).value

        # if the field is not blank, then
        if DateInSeconds:
            print ("Row: " + str(iRowIndexNumber) + " Raw Seconds: " + str(DateInSeconds) + " Days: " + str(DateInSeconds/28800) )
            WorksheetName.cell(iRowIndexNumber, kDateColumn).value = DateInSeconds/28800
    return
