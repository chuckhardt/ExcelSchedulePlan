# ******************************************************
#
# Southeastern Railway Museum, Project Task Scheduler
#
# Use Python 3.5+ or better
#
# Author/Owner: C. Hardt
#
# ******************************************************

import openpyxl
import datetime
import array as arr
import sys

from openpyxl.styles import NamedStyle
from ScheduleExcelUtilites import ExcelColumnUtil
from ScheduleExcelDateFormat import CalculateNextWorkingDay


# ***************************************
#
# ScheduleExcelCalculateTaskStartAndEndDates()
#
# This is executed only on a given WorkBlk within a worksheet
# It will only be execute if the WorkBlk numbers match
# The start date will be the first day of the WorkBlk
#
# ***************************************
def ScheduleExcelCalculateTaskStartAndEndDates(WorksheetName, sWorkBlkName, sWorkBlkStartDate):

    ExcelColumnUtilities = ExcelColumnUtil()

    kWorkBlkColumnNumber = ExcelColumnUtilities.ExcelFindColumnHeader(WorksheetName, "TASK", 7)
    kAssigneeNameColumnNumber = ExcelColumnUtilities.ExcelFindColumnHeader(WorksheetName, "Assignee", 7)
    kEstimatedTaskTimeInDaysColumnNumber = ExcelColumnUtilities.ExcelFindColumnHeader(WorksheetName, "WORK DAYS", 7)
    kEstimatedTaskTimeInSecondsColumnNumber = ExcelColumnUtilities.ExcelFindColumnHeader(WorksheetName, "SECONDS", 7)
    kStartDateColumnNumber = ExcelColumnUtilities.ExcelFindColumnHeader(WorksheetName, "START", 7)
    kEndDateColumnNumber = ExcelColumnUtilities.ExcelFindColumnHeader(WorksheetName, "END", 7)

    # Check and make sure we found all the columns we were looking for
    if ((kAssigneeNameColumnNumber == -1) or (kEstimatedTaskTimeInSecondsColumnNumber == -1) or
        (kStartDateColumnNumber == -1) or (kEndDateColumnNumber == -1) or
        (kWorkBlkColumnNumber == -1) or (kEstimatedTaskTimeInDaysColumnNumber == -1)):
        print ("ERROR: Column(s) not found")
        return

    kMonday = 1
    kFriday = 5

    try:

        bWorkBlkNameFoundFlag = False

        # we need to convert the string date into numbers
        iCalendarYear = int(sWorkBlkStartDate[6:10])
        iCalendarMonth = int(sWorkBlkStartDate[:2])
        iCalendarDay = int(sWorkBlkStartDate[3:5])

        FirstDayOfWorkBlk = datetime.datetime(iCalendarYear, iCalendarMonth, iCalendarDay, 00, 00)
        MidNightToday = datetime.datetime(iCalendarYear, iCalendarMonth, iCalendarDay, 00, 00)

        sLastAssignee = ""
        iMaxRowNumber = WorksheetName.max_row
        iAssigneeTotalNumberOfWorkingSecondsSoFarToday = 0
        DaysFlag = False
        OverflowTimeWhichHasToBeTransferedToTheNextTask = 0

        MyDateCellStyle = NamedStyle(name='mydatetime', number_format='MM/DD/YYYY HH:MM')

        # we are going to process the file one line at a time.
        for iRowNumber in range(2, iMaxRowNumber + 1, 1):

            # We need to find the row that contains the WorkBlk number we are looking for, When we 
            # find it, we will start.  After finding the right row, we will stop when we encounter
            # a blank line or the word WorkBlk
            if (bWorkBlkNameFoundFlag == True):
                if (WorksheetName.cell(iRowNumber, 4).value != ""):
                    print ("Ending the Parse as we Found: " + WorksheetName.cell(iRowNumber, 4).value)
                    return
        
            else:    
                sWorkBlkNumber = WorksheetName.cell(iRowNumber, kWorkBlkColumnNumber).value
                if (sWorkBlkNumber != sWorkBlkName):
                    continue
                # we found the matching WorkBlk number, continue processing the WorkBlk until we 
                # encounter a blank line or the next block name
                else:      
                    bWorkBlkNameFoundFlag = True  # we found at least one row with the correct (searched for) WorkBlk name

                    
            # Get Assignee Name, if different then the last name, then save it
            sAssigneeName =  WorksheetName.cell(iRowNumber,kAssigneeNameColumnNumber).value
            if (sAssigneeName != sLastAssignee):
                print ("New Assignee: " + sAssigneeName)
                sLastAssignee = sAssigneeName
                iAssigneeTotalNumberOfWorkingSecondsSoFarToday = 0
                OverflowTimeWhichHasToBeTransferedToTheNextTask = 0
                DaysFlag = False

                # If we have a new assignee, we are going to use the WorkBlk start date as the new start
                ExcelTaskStartDate = FirstDayOfWorkBlk

            else:
                ExcelTaskStartDate = ExcelCalculatedTaskCompletionDate


            # We are going to check the day of the week.  If Sat or Sun, then increment the day
            ExcelTaskStartDate = CalculateNextWorkingDay(ExcelTaskStartDate)

            # write the start date of the task back to spreadsheet
            WorksheetName.cell(iRowNumber, kStartDateColumnNumber).value = ExcelTaskStartDate

            # Get the time estimate to complete the task/defect
            ExcelTaskOriginalEstimate = WorksheetName.cell(iRowNumber, kEstimatedTaskTimeInSecondsColumnNumber).value

            ExcelTaskOriginalEstimate += OverflowTimeWhichHasToBeTransferedToTheNextTask

            # convert the time estimate (which is in seconds), into days.  The remainder is in seconds
            OriginalEstimateConvertedToDays = ExcelTaskOriginalEstimate // 28800
            OriginalEstimateConvertedToSeconds = ExcelTaskOriginalEstimate % 28800

            #print(" ReCal Days/Secs: " + str(OriginalEstimateConvertedToDays) + "/" + str(OriginalEstimateConvertedToSeconds), end = " ")

            #####
            # Now we need to determine if we are exceeding 8 hours in one day.  If we are, then we have to increment the
            # number of days.
            if ((OriginalEstimateConvertedToDays == 0) and (OriginalEstimateConvertedToSeconds > 0)):

                if ((iAssigneeTotalNumberOfWorkingSecondsSoFarToday + OriginalEstimateConvertedToSeconds) >= 28800):

                    TodaysMaxRemainingDailyBalance = 28800 - iAssigneeTotalNumberOfWorkingSecondsSoFarToday
                    iAssigneeTotalNumberOfWorkingSecondsSoFarToday += TodaysMaxRemainingDailyBalance

                    OverflowTimeWhichHasToBeTransferedToTheNextTask = OriginalEstimateConvertedToSeconds - TodaysMaxRemainingDailyBalance

                    if (OverflowTimeWhichHasToBeTransferedToTheNextTask > 28800):
                        OverflowTimeWhichHasToBeTransferedToTheNextTask = 28800 - (iAssigneeTotalNumberOfWorkingSecondsSoFarToday + OriginalEstimateConvertedToSeconds) - TodaysMaxRemainingDailyBalance
                    #else:
                    #    OverflowTimeWhichHasToBeTransferedToTheNextTask = 0
                         #print (" Xfer: " + str(OverflowTimeWhichHasToBeTransferedToTheNextTask), end = " ")

                    DaysFlag = True

                    OriginalEstimateConvertedToSeconds = TodaysMaxRemainingDailyBalance

                    #print ("Overflow, Sum 4 Day: " + str(iAssigneeTotalNumberOfWorkingSecondsSoFarToday), end = " ")

                else:
                    iAssigneeTotalNumberOfWorkingSecondsSoFarToday += OriginalEstimateConvertedToSeconds
                    #print("Underflow, Assignee Sum 4 Day: " + str(iAssigneeTotalNumberOfWorkingSecondsSoFarToday), end = " ")
                    OverflowTimeWhichHasToBeTransferedToTheNextTask = 0

            # now format the cell into the date time format
            #WorksheetName.cell(iRowNumber,kEndDateColumnNumber).style = MyDateCellStyle

            ExcelCalculatedTaskCompletionDate = ExcelTaskStartDate + datetime.timedelta(days=OriginalEstimateConvertedToDays, seconds=OriginalEstimateConvertedToSeconds)
            #print ("End Date: " + str(ExcelCalculatedTaskCompletionDate), end = " ")

            # We are going to check the day of the week.  If Sat or Sun, then increment the day
            ExcelCalculatedTaskCompletionDate = CalculateNextWorkingDay(ExcelCalculatedTaskCompletionDate)
            #print("ReCal End Date: " + str(ExcelCalculatedTaskCompletionDate))

            WorksheetName.cell(iRowNumber, kEndDateColumnNumber).value = ExcelCalculatedTaskCompletionDate

            # now we need to convert the estimated time in seconds to days.
            WorksheetName.cell(iRowNumber, kEstimatedTaskTimeInDaysColumnNumber).value = (WorksheetName.cell(iRowNumber, kEstimatedTaskTimeInSecondsColumnNumber).value /28800)


            if (DaysFlag == True):
                OriginalEstimateConvertedToDays += 1
                DaysFlag = False
                ExcelCalculatedTaskCompletionDate = MidNightToday + datetime.timedelta(days=1)
                MidNightToday = MidNightToday + datetime.timedelta(days=1)
                iAssigneeTotalNumberOfWorkingSecondsSoFarToday = 0

        if (bWorkBlkNameFoundFlag == False):
            print ("ERROR the WorkBlk " + sWorkBlkName + " NOT FOUND!")

    except Exception as e:
        print("Exception - ScheduleExcelCalculateTaskStartAndEndDates() " + str(e))

